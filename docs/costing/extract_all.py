"""
Parse all .xlsx files in docs/costing and produce extracted_all_costing.json.

Output format (app.py loads this directly at startup):
  { "<capacity>": { "<config_key>": { ...component fields... }, ... }, ... }

Run this script whenever the Excel files change:
  python docs/costing/extract_all.py
Then restart the Flask app; no code changes needed.
"""
import sys, re, json
sys.path = [p for p in sys.path if 'costing' not in p.lower()]
import openpyxl
from pathlib import Path

COSTING_DIR = Path(__file__).parent
CAPS = ['6', '8', '10', '13', '15', '20', '26']

# Column index (0-based) for unit_cost and total per config
COL_UNIT  = {'MS Auto Door':6,'SS Auto Door':8,'MS Manual Door':10,'SS Manual Door':12,'MS Gearless Auto Door':14,'SS Gearless Auto Door':16}
COL_TOTAL = {'MS Auto Door':7,'SS Auto Door':9,'MS Manual Door':11,'SS Manual Door':13,'MS Gearless Auto Door':15,'SS Gearless Auto Door':17}

# In golden/rose-gold files the MS column = Golden, SS column = Rose Gold
MS_COL_TO_KEY = {
    'MS Auto Door':          'MS Auto Door',
    'SS Auto Door':          'SS Auto Door',
    'MS Manual Door':        'MS Manual Door',
    'SS Manual Door':        'SS Manual Door',
    'MS Gearless Auto Door': 'MS Gearless Auto Door',
    'SS Gearless Auto Door': 'SS Gearless Auto Door',
}
RG_COL_TO_KEY = {
    'MS Auto Door':          'Golden Auto Door',
    'SS Auto Door':          'Rose Gold Auto Door',
    'MS Manual Door':        'Golden Manual Door',
    'SS Manual Door':        'Rose Gold Manual Door',
    'MS Gearless Auto Door': 'Golden Gearless Auto Door',
    'SS Gearless Auto Door': 'Rose Gold Gearless Auto Door',
}

GEARED_ROPES   = {'6':3,'8':4,'10':4,'13':4,'15':4,'20':5,'26':5}
GEARLESS_ROPES = {'6':4,'8':4,'10':5,'13':5,'15':5,'20':8,'26':8}


def classify_file(name):
    n = name.lower()
    m = re.search(r'(\d+)p', n)
    cap = int(m.group(1)) if m else None
    vision = 'SV' if 'small' in n else 'BV'
    finish = 'RG' if ('golden' in n or 'rose' in n) else 'MS'
    return vision, finish, cap


def get_val(row, col):
    if col >= len(row):
        return None
    v = row[col]
    return float(v) if isinstance(v, (int, float)) and v > 0 else None


def parse_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    hdr = next((i for i, r in enumerate(rows) if str(r[0] or '').strip() == 'S.No.'), None)
    if hdr is None:
        return {}
    items = {}
    for row in rows[hdr + 1:]:
        name = str(row[1] or '').strip()
        if not name:
            continue
        entry = {'qty': row[4], 'actual': row[5], 'unit_costs': {}, 'totals': {}}
        for cfg, col in COL_UNIT.items():
            entry['unit_costs'][cfg] = get_val(row, col)
        for cfg, col in COL_TOTAL.items():
            entry['totals'][cfg] = get_val(row, col)
        if any(v is not None for v in entry['unit_costs'].values()):
            items.setdefault(name, []).append(entry)
    return items


def fu(items, name, cfg, fb=0):
    """First unit cost for item+config."""
    for row in items.get(name, []):
        v = row['unit_costs'].get(cfg)
        if isinstance(v, (int, float)) and v > 0:
            return v
    return fb


def ft(items, name, cfg, fb=0):
    """First total for item+config."""
    for row in items.get(name, []):
        v = row['totals'].get(cfg)
        if isinstance(v, (int, float)) and v > 0:
            return v
    return fb


def extract_config(items, col_cfg, cap_str, gearless):
    ctrl = fu(items, 'Controlller with Drive +DBR', col_cfg) or fu(items, 'Controller with Drive +DBR', col_cfg)

    brk_rows = items.get('Bracket', [])
    if gearless:
        brk_val = (brk_rows[1]['unit_costs'].get(col_cfg) if len(brk_rows) > 1 else None)
        brk_rate = brk_val if isinstance(brk_val, (int, float)) and brk_val > 0 else 1350
    else:
        brk_rate = fu(items, 'Bracket', col_cfg, 950)

    cable_rows = items.get('Cable', [])
    cable = 12236.8
    if cable_rows:
        cr = cable_rows[0]
        cq = float(cr.get('actual') or cr.get('qty') or 0)
        cu = cr['unit_costs'].get(col_cfg) or 0
        if cq > 0 and cu > 0:
            cable = round(cq * cu, 2)

    wrows = items.get('Wiring', [])[:5]
    wt = []
    for wr in wrows:
        t = wr['totals'].get(col_cfg)
        wt.append(int(t) if isinstance(t, (int, float)) and t > 0 else 0)
    while len(wt) < 5:
        wt.append(0)

    rope_rate, rope_base = 0, 0.0
    for rr in items.get('Rope', []):
        u = rr['unit_costs'].get(col_cfg)
        q = rr.get('actual') or rr.get('qty')
        if isinstance(u, (int, float)) and u > 0:
            rope_rate = int(u)
            rope_base = float(q) if isinstance(q, (int, float)) else 0.0
            break

    load = fu(items, 'loading & Unloading', col_cfg) or ft(items, 'loading & Unloading', col_cfg) or 5000
    scaff = fu(items, 'Scafolding', col_cfg) or ft(items, 'Scafolding', col_cfg) or 7500

    return {
        'car_cabin':         int(fu(items, 'Car cabin', col_cfg, 45000)),
        'overload':          int(fu(items, 'Overload', col_cfg, 4000)),
        'cabin_packing':     int(fu(items, 'Cabin Packing Charges', col_cfg, 2500)),
        'cabin_freight':     int(fu(items, 'Cabin inward Transport & local Freight', col_cfg, 6000)),
        'controller':        int(ctrl or 0),
        'ard_battery':       int(fu(items, 'ARD (UPS) with battery', col_cfg, 0)),
        'motor':             int(fu(items, 'Motor', col_cfg, 0)),
        'motor_hosting':     int(fu(items, 'Motor Hosting', col_cfg, 2500)),
        'safety':            int(fu(items, 'Geared/Gearless Safety', col_cfg, 0)),
        'osg_rope':          round(ft(items, 'OSG with rope', col_cfg, 0), 2),
        'weight_counter':    round(fu(items, 'Weight counter with granite floor', col_cfg, 0), 2),
        'sensor_door':       int(fu(items, 'Sensor door', col_cfg, 0)),
        'car_door':          int(fu(items, 'Car Door', col_cfg, 0)),
        'landing_door_rate': int(fu(items, 'Landing Door', col_cfg, 0)),
        'cable':             cable,
        'wiring_totals':     wt,
        'other':             int(fu(items, 'OTHER', col_cfg, 25000)),
        'freight':           int(fu(items, 'Freight', col_cfg, 5000)),
        'loading_unloading': int(load),
        'scaffolding':       int(scaff),
        'guide_rail_rate':   int(fu(items, 'Guide rail', col_cfg, 5000)),
        'bracket_rate':      int(brk_rate),
        'rope_rate':         rope_rate,
        'rope_num_ropes':    GEARLESS_ROPES[cap_str] if gearless else GEARED_ROPES[cap_str],
    }


def build_variants():
    file_map = {}
    for f in COSTING_DIR.glob('*.xlsx'):
        vision, finish, cap = classify_file(f.name)
        if cap is not None:
            file_map[(vision, finish, cap)] = f

    variants = {}
    for (vision, finish, cap), fpath in sorted(file_map.items()):
        vk = f'{vision}_{finish}'
        items = parse_file(fpath)
        cap_str = str(cap)
        col_map = RG_COL_TO_KEY if finish == 'RG' else MS_COL_TO_KEY
        if vk not in variants:
            variants[vk] = {}
        variants[vk][cap_str] = {}
        for col_cfg, dest_key in col_map.items():
            gearless = 'Gearless' in col_cfg
            variants[vk][cap_str][dest_key] = extract_config(items, col_cfg, cap_str, gearless)
    return variants


def build_flat(variants: dict) -> dict:
    """Flatten variant dict to {capacity: {config_key: components}}."""
    flat: dict = {}
    for cap in CAPS:
        flat[cap] = {}
        for k, v in variants.get('BV_MS', {}).get(cap, {}).items():
            flat[cap][k] = v                  # Big Vision MS/SS (existing key names)
        for k, v in variants.get('BV_RG', {}).get(cap, {}).items():
            flat[cap][k] = v                  # Big Vision Golden / Rose Gold
        for k, v in variants.get('SV_MS', {}).get(cap, {}).items():
            flat[cap][f'SV {k}'] = v          # Small Vision prefix
    return flat


def main():
    variants = build_variants()
    flat = build_flat(variants)

    out = COSTING_DIR / 'extracted_all_costing.json'
    out.write_text(json.dumps(flat, indent=2), encoding='utf-8')

    total = sum(len(v) for v in flat.values())
    print(f'Saved {out}')
    print(f'  {len(flat)} capacities, {total} config entries total')
    for cap in CAPS:
        print(f'  {cap}p: {list(flat[cap].keys())}')


if __name__ == '__main__':
    main()
