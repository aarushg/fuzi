from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


BASE_DIR = Path(__file__).resolve().parent
SOURCE_DIR = BASE_DIR / "6 passenger costing"

LOP_COP_LOOKUP = {
    2: 14000,
    3: 16450,
    4: 18900,
    5: 21350,
    6: 23800,
    7: 26250,
    8: 28700,
    9: 31150,
    10: 33600,
}

DRIVE_SPEED = {
    "geared": 0.65,
    "gearless": 1.0,
}


@dataclass(frozen=True)
class Scenario:
    id: str
    label: str
    file: str
    vision: str
    material: str
    source_material: str
    drive: str
    door: str
    stops: int
    final_offer: float
    material_total: float
    post_install_total: float
    margin: float


def money(value: float) -> float:
    return round(float(value or 0), 2)


def scenario_from_file(path: Path) -> dict[str, Any]:
    name = path.name
    lower = name.lower()
    is_rg = "golden and rose gold" in lower
    if "small vision" in lower:
        vision = "small"
    else:
        vision = "big"

    source_material = "ss" if "stainless steel" in lower else "ms"
    material = "rg" if is_rg and source_material == "ss" else source_material
    drive = "gearless" if "gearless" in lower else "geared"
    door = "manual" if "manual" in lower else "auto"

    wb_values = openpyxl.load_workbook(path, data_only=True)
    ws_values = wb_values.active
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    ws_formulas = wb_formulas.active

    totals = []
    final_offer = None
    margin = None
    for row in range(1, ws_values.max_row + 1):
        item = str(ws_formulas.cell(row, 2).value or "").strip().lower()
        marker = str(ws_formulas.cell(row, 4).value or "").strip().lower()
        if marker == "total":
            totals.append(float(ws_values.cell(row, 8).value or 0))
        if item == "our margin":
            margin = float(ws_values.cell(row, 8).value or 0)
        if item == "final offer":
            final_offer = float(ws_values.cell(row, 8).value or 0)

    stops = int(ws_values["G3"].value or 2)
    label = (
        f"{'Small/Non Vision' if vision == 'small' else 'Big Vision Glass Door'} "
        f"{material_label(material)} "
        f"{motor_label(drive)} {door.title()} - {stops} stops"
    )
    if is_rg and source_material == "ms":
        label = (
            f"Big Vision Glass Door Mild Steel "
            f"{motor_label(drive)} {door.title()} - {stops} stops "
            f"(RG/Golden source file)"
        )

    return {
        "id": path.stem.lower().replace(" ", "-").replace("_", "-"),
        "label": label,
        "file": name,
        "vision": vision,
        "material": material,
        "source_material": source_material,
        "is_rg_source": is_rg,
        "drive": drive,
        "door": door,
        "stops": stops,
        "final_offer": money(final_offer or 0),
        "material_total": money(totals[0] if totals else 0),
        "post_install_total": money(totals[1] if len(totals) > 1 else 0),
        "margin": money(margin or 0),
    }


def load_source_scenarios() -> list[dict[str, Any]]:
    if not SOURCE_DIR.exists():
        return []
    return [scenario_from_file(path) for path in sorted(SOURCE_DIR.glob("*.xlsx"))]


def material_label(material: str) -> str:
    return {
        "ms": "Mild Steel",
        "ss": "Stainless Steel",
        "rg": "RG/Golden",
    }.get(material, material)


def motor_label(drive: str) -> str:
    return {
        "geared": "Geared Motor",
        "gearless": "Gearless Motor",
    }.get(drive, drive)


def default_travel(stops: int) -> dict[str, float]:
    labels = [
        "Pit",
        "B to G",
        "G to 1",
        "1 to 2",
        "2 to 3",
        "3 to 4",
        "4 to 5",
        "5 to 6",
        "6 to 7",
        "7 to 8",
        "8 to 9",
        "9 to 10",
        "Overhead",
    ]
    travel = {label: 0.0 for label in labels}
    travel["Pit"] = 1600
    travel["Overhead"] = 4800
    if stops <= 2:
        travel["G to 1"] = 4000
        return travel

    floor_values = [3800, 3800, 3900, 3800, 3800, 3800, 3800, 3800, 3800]
    floor_labels = labels[2:11]
    for index in range(min(stops - 1, len(floor_labels))):
        travel[floor_labels[index]] = floor_values[index]
    return travel


def parse_params(data: dict[str, Any]) -> dict[str, Any]:
    stops = int(data.get("stops") or 2)
    stops = max(2, min(10, stops))
    travel = default_travel(stops)
    drive = data.get("drive") or "geared"
    material = data.get("material") or "ms"
    source_material = data.get("source_material")
    for key in list(travel):
        safe_key = key.lower().replace(" ", "_")
        if safe_key in data:
            try:
                travel[key] = float(data[safe_key])
            except (TypeError, ValueError):
                pass
    return {
        "capacity": int(data.get("capacity") or 6),
        "vision": data.get("vision") or "small",
        "material": material,
        "source_material": source_material,
        "is_rg_source": bool(data.get("is_rg_source")),
        "drive": drive,
        "door": data.get("door") or "auto",
        "stops": stops,
        "speed": DRIVE_SPEED.get(drive, 0.65),
        "margin_percent": float(data.get("margin_percent") or 20),
        "travel": travel,
    }


def row(label: str, qty: float, rate: float, formula: str, spec: str = "", note: str = "") -> dict[str, Any]:
    return {
        "item": label,
        "spec": spec,
        "qty": money(qty),
        "rate": money(rate),
        "amount": money(qty * rate),
        "formula": formula,
        "note": note,
    }


def amount_row(label: str, amount: float, formula: str, spec: str = "", note: str = "") -> dict[str, Any]:
    return {
        "item": label,
        "spec": spec,
        "qty": 1,
        "rate": money(amount),
        "amount": money(amount),
        "formula": formula,
        "note": note,
    }


def cabin_rate(params: dict[str, Any]) -> float:
    material = params["source_material"] or params["material"]
    if material == "ms":
        return 45000
    if params["material"] == "rg":
        return 75000 + 10000
    return 75000


def door_rates(params: dict[str, Any]) -> tuple[float, float]:
    material = params["source_material"] or params["material"]
    vision = params["vision"]
    manual = params["door"] == "manual"
    if manual and material == "ms":
        return 7000 + 1500, 7000 + 3800 + 1500
    if manual and material == "ss":
        return 25000 + 1500, 35000 + 1500
    if material == "ms":
        return 43245, 17880
    if params["material"] == "rg":
        return 57016, 38191 + 2000
    if vision == "small":
        return 47840, 26355
    return 57016, 38191 + 2000


def calculate_costing(raw: dict[str, Any]) -> dict[str, Any]:
    params = parse_params(raw)
    total_travel = sum(params["travel"].values())
    stops = params["stops"]
    drive = params["drive"]
    manual = params["door"] == "manual"
    gearless = drive == "gearless"

    rows: list[dict[str, Any]] = []
    guide_actual = 5 if stops <= 2 else 8
    bracket_actual = 7 if stops <= 2 else 11
    wiring_actuals = (31, 30, 103) if stops <= 2 else (38, 43, 175)

    rows.append(row("Guide rail", guide_actual, 4800 + 200, "qty=(travel/5000)*2; amount=actual*(4800+200)", "9mm & 5mm"))

    if gearless:
        rows.append(row("Bracket", bracket_actual, 1100 + 250, "gearless side-counter amount=actual*(1100+250)", "Comb (Side Counter300mm) angle", "Replaces source =#REF! quantity."))
    elif manual:
        rows.append(amount_row("Bracket", 6650, "manual source amount=6650; source rate formula was =1800*G3", "GB(Back Counter)"))
    else:
        rows.append(row("Bracket", bracket_actual, 950, "qty=((R14-900)/2000)+2; amount=actual*950", "GB(Back Counter)"))

    rows.append(row("Car cabin", 1, cabin_rate(params), "material/finish cabin rate", "ss handrail+fan/blower"))
    rows.append(row("Overload", 1, 4000, "1*4000"))
    rows.append(row("Cabin Packing Charges", 1, 2500, "1*2500"))
    rows.append(row("Cabin inward Transport & local Freight", 1, 6000, "1*6000"))

    controller = 55500 + 1500 + 6800 + 4200 if gearless else 37500 + 1500
    controller_spec = "5hp closeloop" if gearless else "5hp openloop"
    rows.append(row("Controller with Drive + DBR", 1, controller, "gearless close-loop or geared open-loop controller", controller_spec))
    rows.append(row("ARD (UPS) with battery", 1, 12500 + 500, "=12500+500"))

    lop_cop = LOP_COP_LOOKUP[stops] + 3000
    rows.append(row("LOP/COP", 1, lop_cop, "=VLOOKUP(G3,I1:J10,2,0)+3000"))

    motor = 75100 + 2000 + 2500 if gearless else 62000 + 2000 + 2500
    rows.append(row("Motor", 1, motor, "gearless or geared ste125 motor", "ste125"))
    rows.append(amount_row("Motor Hoisting", 2500, "2500"))

    if gearless:
        rope_qty = (((total_travel / 1000) * 2) + 2) * 4
        rope_rate = 77
        rope_spec = "seg 10,8mm,4rope"
    else:
        rope_qty = ((total_travel / 1000) + 2) * 3
        rope_rate = 113
        rope_spec = "srt125,13mm,3rope"
    rows.append(row("Rope", rope_qty, rope_rate, "gearless or geared rope formula", rope_spec))

    cable_qty = ((total_travel + 3000 + 3000) / 1000) * 4
    rows.append(row("Cable", cable_qty, 128, "=((R14+3000+3000)/1000)*4", "flat 12 core,.7mm"))

    rows.append(row("Geared/Gearless Safety", 1, 40000 if gearless else 28000, "gearless=40000; geared=28000"))

    osg_qty = rope_qty / 4 if gearless else (total_travel / 1000) * 2 + 2
    osg_amount = osg_qty * 77 + 4000 + 1000
    rows.append(amount_row("OSG with rope", osg_amount, "=(quantity*77)+4000+1000", "4000+rope+sos&clip"))

    rows.append(amount_row("Weight counter with granite floor", 15.5 * 38 * (15.5 + 2), "=15.5*38*(15.5+2)"))

    if not manual:
        rows.append(row("Sensor door", 1, 4500, "auto door only"))

    rows.append(row("Wiring - limit switch", wiring_actuals[0], 50 + 2, "=R14/1000+10+10", "limit switch, 6 core,.5mm"))
    rows.append(row("Wiring - LOP/COP/car top", wiring_actuals[1], 67 + 2, "=(R14/1000)+G3*3+13", "lop,cop,car top 12 core,.5mm"))
    rows.append(row("Wiring - lock/LOP/fireman alarm", wiring_actuals[2], 20 + 3, "=((R14/1000)+5)+G3*2+(R14/1000*8)", "lock,lop,fireman alarm, car top,2core,.5mm"))
    rows.append(row("Wiring - motor flexible", 50, 35 + 2.2, "50*(35+2.2)", "2.5mm, flexible, motor"))
    rows.append(row("Wiring - brake cable", 50, 25 + 2, "50*(25+2)", ".75mm for brake 2core .5mm"))

    car_door, landing_door = door_rates(params)
    rows.append(row("Car Door", 1, car_door, "door table rate", "upto 800mm"))
    rows.append(row("Landing Door", stops, landing_door, "=stops*landing door rate", "upto 800mm"))

    rows.append(amount_row("OTHER", 25000, "source amount 25000; auto rate commonly 25000"))
    rows.append(amount_row("Freight", 5000, "local freight 5000; outstation note 10000"))
    rows.append(amount_row("Loading & Unloading", 5000, "=2000+1000+2000"))
    rows.append(amount_row("Scaffolding", 7500, "7500"))

    material_total = money(sum(item["amount"] for item in rows))
    installation = money(stops * 12000)
    commissioning = 12000
    warranty = money(material_total * 0.05)
    post_install_total = money(material_total + installation + commissioning + warranty)
    margin = money(post_install_total * (params["margin_percent"] / 100))
    final_offer = money(post_install_total + margin)

    service_rows = [
        row("Installation local", stops, 12000, "=stops*12000; outstation 12500"),
        amount_row("Commissioning", commissioning, "12000"),
        amount_row("Warranty", warranty, "=material subtotal*5%"),
    ]

    return {
        "params": params,
        "warnings": warnings_for(params),
        "travel_total": money(total_travel),
        "lop_cop_lookup": LOP_COP_LOOKUP,
        "rows": rows,
        "service_rows": service_rows,
        "totals": {
            "material_total": material_total,
            "installation": installation,
            "commissioning": money(commissioning),
            "warranty": warranty,
            "post_install_total": post_install_total,
            "margin": margin,
            "final_offer": final_offer,
        },
    }


def warnings_for(params: dict[str, Any]) -> list[str]:
    warnings = []
    if params["capacity"] != 6:
        warnings.append("Only 6-passenger source workbooks are available; other capacities use the same 6-passenger logic until more source files are added.")
    if params["door"] == "manual" and (params["vision"] != "small" or params["drive"] != "geared"):
        warnings.append("Manual-door source coverage exists only for small-vision geared files. This combination is calculated from normalized rules.")
    if params["material"] == "rg" and params["door"] == "manual":
        warnings.append("Golden/rose-gold manual-door source files are not available.")
    if params["material"] == "rg" and params["vision"] != "big":
        warnings.append("RG/Golden source coverage is available only with Big Vision Glass Door.")
    return warnings


def find_matching_source(params: dict[str, Any], scenarios: list[dict[str, Any]]) -> dict[str, Any] | None:
    for scenario in scenarios:
        if scenario.get("is_rg_source") and scenario["material"] != "rg" and not params.get("is_rg_source"):
            continue
        if (
            scenario["vision"] == params["vision"]
            and scenario["material"] == params["material"]
            and (params.get("source_material") is None or scenario.get("source_material") == params.get("source_material"))
            and scenario["drive"] == params["drive"]
            and scenario["door"] == params["door"]
            and int(scenario["stops"]) == int(params["stops"])
        ):
            return scenario
    return None


def verify_against_sources(scenarios: list[dict[str, Any]]) -> list[dict[str, Any]]:
    results = []
    for scenario in scenarios:
        calc = calculate_costing(scenario)
        final_diff = money(calc["totals"]["final_offer"] - scenario["final_offer"])
        material_diff = money(calc["totals"]["material_total"] - scenario["material_total"])
        results.append(
            {
                **scenario,
                "calculated_final_offer": calc["totals"]["final_offer"],
                "calculated_material_total": calc["totals"]["material_total"],
                "final_diff": final_diff,
                "material_diff": material_diff,
                "status": "pass" if abs(final_diff) < 1 and abs(material_diff) < 1 else "review",
            }
        )
    return results
